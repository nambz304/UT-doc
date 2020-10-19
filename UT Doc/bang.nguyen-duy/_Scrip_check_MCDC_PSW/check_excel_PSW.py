import sys
import openpyxl
import win32com.client
from win32com.client import Dispatch
# from datetime import date


# '''variable const'''
name_user = 'EXTERNAL Nguyen Duy Bang (Ban Vien, RBVH/EPS45)'

color_Red = '\x1b[91m '
color_Green = '\x1b[92m '
color_Yellow = '\x1b[93m '
color_Orange = '\x1b[38;5;208m '
color_White = '\x1b[0m '
E_OK = True
E_NOT_OK = False

def print_error(string=None, col_checking=None):
    """
    Print red color
    col_checking is type int.
    """
    if col_checking is not None:
        print(color_Red + string + color_White + str(sheet_TCs.cell(Row_Name_Var, col_checking).value))
    else:
        print(color_Red + string + color_White)
    return


def print_notice(string=None, col_checking=None):
    """
    Print green color
    col_checking is type int.
    """
    if col_checking is not None:
        print(color_Green + string + color_White + str(sheet_TCs.cell(Row_Name_Var, col_checking).value))
    else:
        print(color_Green + string + color_White)
    return


def print_warning(string=None, col_checking=None):
    """
    Print yellow color
    col_checking is type int.
    """
    if col_checking is not None:
        print(color_Yellow + string + color_White + str(sheet_TCs.cell(Row_Name_Var, col_checking).value))
    else:
        print(color_Yellow + string + color_White)
    return

# """ Lay input tu bash shell """
path_excel = str(sys.argv[1]) # link file TD
value_C0   = str(sys.argv[2]) # input chua data C0 RTRT
value_C1   = str(sys.argv[3]) # input chua data C1 RTRT
value_MCDC = str(sys.argv[4]) # input chua data MCDC RTRT

try:
    xl = win32com.client.Dispatch('Excel.Application')
    print(path_excel)
    wb = xl.Workbooks.Open(path_excel)
    readData  = wb.WorkSheets('Unit_Test_Info')
    writeData = wb.Worksheets('Unit_Test_Info')
    allData   = readData.UsedRange 

    # Get number of rows used on active sheet
    getNumRows = allData.Rows.Count
    # print ('Number of rows used in sheet : ',getNumRows)

    #Get number of columns used on active sheet
    getNumCols = allData.Columns.Count
    # print ('Number of columns used in sheet : ',getNumCols)
except:
    print_error(" ====> Error: Excel")
    wb.Close()
    xl.Quit()
    exit() # Exit khoi python



def read_cell(row=0, col=0):
    try:
        return allData.Cells(row,col).value
    except:
        print_error("Read failed.")

def write_cell(row=0, col=0, msg=None):
    try:
        col = col +1
        writeData.Cells(row,col).Value = msg
    except:
        print_error("Write failed.")

for row in range(1, getNumRows +1):
    for col in range(1, getNumCols +1):
        if read_cell(row,col) == "Tester":
            row_tester = row
            col_tester = col
            # Write on empty cell of active sheet
            write_cell(row,col +1,name_user)
            print("Tester: ",read_cell(row,col +1))
        elif read_cell(row,col) == "Date":
            row_date = row
            col_date = col
            # today = date.today()
            # d1 = today.strftime("%m/%d/%Y")
            # write_cell(row, col +1, d1)
            print("Date: ", read_cell(row,col +1))
        elif read_cell(row,col) == "Item_Name":
            row_Item_Name = row
            col_Item_Name = col
            print("Item_Name: ", read_cell(row,col +1))
        elif read_cell(row,col) == "C0":
            row_C0 = row
            col_C0 = col
            write_cell(row, col +1, value_C0)
            print(read_cell(row,col +1))
            if float(read_cell(row,col +1)) < 100:
                print("aaaaaaaaaaaaaaaaaaaa11111")
        elif read_cell(row,col) == "C1":
            row_C1 = row
            col_C1 = col
            write_cell(row, col +1, value_C1)
            print(read_cell(row,col +1))
            if float(read_cell(row,col +1)) < 100:
                print("aaaaaaaaaaaaaaaaaaaa2222")
        elif read_cell(row,col) == "MCDC":
            row_MCDC = row
            col_MCDC = col
            write_cell(row, col +1, value_MCDC)
            print(read_cell(row,col +1))
            if float(read_cell(row,col +1)) < 100:
                print("aaaaaaaaaaaaaaaaaaaa333")



# Save excel doc
wb.Save()
# Save As current excel doc
#wb.SaveAs('updatedSample.xlsx')

wb.Close()
xl.Quit()
xl = None