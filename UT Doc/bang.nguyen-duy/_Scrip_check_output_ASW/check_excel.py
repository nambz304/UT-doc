import sys
import openpyxl

'''variable const'''
name_user = 'EXTERNAL Nguyen Duy Bang (Ban Vien, RBVH/EPS45)'
Sheet_default = 3
Sheet_TC_End = 100
TC_No = 1
Row_Enum = 15
Row_Tolerance = 18
Row_Type = 19
Row_Max = 20
Row_Min = 21
Row_Title = 22
Row_Name_Var = 23
Row_TC1 = 24
color_Red = '\x1b[91m '
color_Green = '\x1b[92m '
color_Yellow = '\x1b[93m '
color_Orange = '\x1b[38;5;208m '
color_White = '\x1b[0m '
E_OK = True
E_NOT_OK = False

#"""
#17/7/2020
#     Updated:    
#         Check enum.
#         Print color.
#     Fixed:
#         Khong can check tolerance cho imported parameter
#         Cover them truong hop cho check_value_row_max_min()
#         Fixed truong hop ngoai le cho file excel [long path, file khong ton tai.]
#18/7/2020
#     Updated:
#         E_OK = True - Condition OK.
#         E_NOT_OK = False - Condition NOK.
#         Compare C0 C1 MCDC RTRT<->Excel TD
#     Fixed:
#         Enum 0 -> 1 thi khong can check mid enum.
#18/7/2020:
#     Updated:
#         Cho phep Check het nhung sheet chua TCs.
#         Group function.
#22/7/2020:
#     Updated:
#         Add check Descriptions.
#     Fixed:
#         Bao error khi file RTRT.txt khong co data.
#24/7/2020:
#     Fixed:
#         String enum name = "None"
#27/7/2020:
#     Fixed:
#         Wrong max value enum => out range max enum.
#         Del line print "None" scan TCs.
#30/7/2020:
#     Fixed:
#         Fixed check Missing Row 'Type'
#31/7/2020:
#     Fixed:
#         Fixed % round(MCDC C0C1)
#05/8/2020:
#     Updated:
#         Fixed check khong can tolerance IMPORTED PARAMETERS
#         Check location LOCAL AS INPUT
# """

# Print 
def print_error(string=None, col_checking=None):
    """
    Print red color
    col_checking is type int.
    """
    if col_checking is not None:
        print(color_Red + string + color_White + str(sheet_TCs.cell(Row_Name_Var, col_checking).value))
    else:
        print(color_Red + string + color_White)
    return


def print_notice(string=None, col_checking=None):
    """
    Print green color
    col_checking is type int.
    """
    if col_checking is not None:
        print(color_Green + string + color_White + str(sheet_TCs.cell(Row_Name_Var, col_checking).value))
    else:
        print(color_Green + string + color_White)
    return


def print_warning(string=None, col_checking=None):
    """
    Print yellow color
    col_checking is type int.
    """
    if col_checking is not None:
        print(color_Yellow + string + color_White + str(sheet_TCs.cell(Row_Name_Var, col_checking).value))
    else:
        print(color_Yellow + string + color_White)
    return
# End # Print 

def check_revision_history():
    """
    Check Fill Date
    """
    if sheet_Rev_History['D17'].value is None:
        print_warning(" ====> WARNING: Revision History: Fill 'Date' ")
# End def check_revision_history():


def check_c0c1_mcdc():
    """
    Check MCDC RTRT <-> TD Excel
    """    
    flag_compare_MCDC = E_OK
    flag_compare_C0 = E_OK
    flag_compare_C1 = E_OK
    # Read data MCDC tu sheet Summary and compare
    # Check MCDC RTRT <-> TD Excel
    value_TD_MCDC = str(sheet_Summary['B6'].value)
    if value_RTRT_MCDC == '100.0%':
        if value_TD_MCDC == '1':
            pass
        else: # failse MCDC
            flag_compare_MCDC = E_NOT_OK
    elif value_RTRT_MCDC == 'NA':
        if value_TD_MCDC == 'NA':
            pass
        else: # failse MCDC
            flag_compare_MCDC = E_NOT_OK
    else: # Neu gia tri la so 0.0% < float < 100.0% 
        # Read data MCDC tu sheet Summary
        value_TD_MCDC = float(sheet_Summary['B6'].value)
        value_TD_MCDC = str(value_TD_MCDC*100)
        value_TD_MCDC = round(float(value_TD_MCDC),1)
        value_TD_MCDC = str(value_TD_MCDC)
        value_TD_MCDC = str(value_TD_MCDC + '%')
        if value_RTRT_MCDC != value_TD_MCDC:
                # failse MCDC
            flag_compare_MCDC = E_NOT_OK

    # Read data C0C1 tu sheet Summary
    value_TD_C0C1 = str(sheet_Summary['A6'].value)
    value_TD_C0 = value_TD_C0C1.split('/')[0] # split data -> C0
    value_TD_C1 = value_TD_C0C1.split('/')[1] # split data -> C1
    # Compare Data C0C1
    # Check C0 RTRT <-> TD Excel
    if value_RTRT_C0 == '100.0%':
        if value_TD_C0 == '100%':
            pass
        else:
            flag_compare_C0 = E_NOT_OK
    elif value_RTRT_C0 != value_TD_C0:
        flag_compare_C0 = E_NOT_OK
        _value_RTRT_C0 = value_RTRT_C0.split('%')[0]
        _value_TD_C0 = value_TD_C0.split('%')[0]
        if round(float(_value_RTRT_C0),1) == round(float(_value_TD_C0),1):
            flag_compare_C0 = E_OK

    # Check C1 RTRT <-> TD Excel
    if value_RTRT_C1 == '100.0%':
        if value_TD_C1 == '100%':
            pass
        else:
            flag_compare_C1 = E_NOT_OK
    elif value_RTRT_C1 != value_TD_C1:
        flag_compare_C1 = E_NOT_OK
        _value_RTRT_C1 = value_RTRT_C1.split('%')[0]
        _value_TD_C1 = value_TD_C1.split('%')[0]
        if round(float(_value_RTRT_C1),1) == round(float(_value_TD_C1),1):
            flag_compare_C1 = E_OK
    # print resuit
    if flag_compare_MCDC == E_NOT_OK:
        print_error(" ====> Error: Failed MCDC. ")
        print_warning("         RTRT/TD_excel: " + value_RTRT_MCDC + "/" +  value_TD_MCDC)
    if flag_compare_C0 == E_NOT_OK:
        print_error(" ====> Error: Failed C0. ")
        print_warning("         RTRT/TD_excel: " + value_RTRT_C0 + "/" +  value_TD_C0)
    if flag_compare_C1 == E_NOT_OK:
        print_error(" ====> Error: Failed C1. ")
        print_warning("         RTRT/TD_excel: " + value_RTRT_C1 + "/" +  value_TD_C1)
    # # check Reason
    # if value_RTRT_C0 != "100.0%"\
    #     or value_RTRT_C1 != "100.0%"\
    #     or value_RTRT_MCDC != "100.0%":
    #     if value_RTRT_MCDC != "NA":
    #         if str(sheet_Summary['C6'].value) == "NA":
    #             print_error(" ====> Error: Reason sheet Summary must not be 'NA'")

    print(" Reason : ", sheet_Summary['C6'].value, "\n")
    #end check c0c1 mcdc
    return
# End def check_c0c1_mcdc():


def check_TM_name():
    """Check TM name"""
    if sheet_TCs.cell(Row_Name_Var, TC_No).value is not None:
        pass
    else:
        print_error(" ====> Error: [A23] Thieu TM_ name at sheet Testcases")
    return
# End def check_TM_name():


def check_sum_TCs():
    """Check xem co bao nhieu TCs"""
    print_notice("===> Tong so TCs: " + str(max_row_table - Row_Name_Var))
    return
# End def check_sum_TCs():


def check_value_row_max_min():
    """check row max min"""
    col_end_row_max = 0
    for col in range(1, max_column_table + 1):
        if str(sheet_TCs.cell(Row_Title, col).value) == 'DESCRIPTIONS':
            col_end_row_max = col
            break
    for col in range(col_start_input, col_end_row_max):
        flag_value_max = E_NOT_OK
        flag_value_min = E_NOT_OK
        if sheet_TCs.cell(Row_Max, col).value is not None:
            flag_value_max = E_OK
        else:
            flag_value_max = E_NOT_OK
        if sheet_TCs.cell(Row_Min, col).value is not None:
            flag_value_min = E_OK
        else:
            flag_value_min = E_NOT_OK

        type_value = str(sheet_TCs.cell(Row_Type, col).value)
        if type_value == 'cont' or type_value == 'log' or type_value == 'enum':
            if flag_value_max and flag_value_min:
                value_max = float(sheet_TCs.cell(Row_Max, col).value)
                value_min = float(sheet_TCs.cell(Row_Min, col).value)
                if value_max < value_min:
                    print_error(" ====> Error: Check Row Max < Min? ", col)
            elif flag_value_max and not flag_value_min:
                print_error(" ====> Error: Missing Row Min: ", col)
            elif not flag_value_max and flag_value_min:
                print_error(" ====> Error: Missing Row Max: ", col)
            else:
                print_error(" ====> Error: Missing Row Max-Min: ", col)
    """End check_value_row_max_min"""
    return
# End def check_value_row_max_min():


def check_input():
    """
    cont
        Check Tolerance YES/NO
        Check must be out max
        Check must be out min
        Check must be mid value
    log
        Check Tolerance 0:1
        Check must be has TRUE & FALSE
        Check fomat "TRUE/FALSE", not 1/0
    enum
        Print warning
        Check Tolerance
    """
    print_notice("======== CHECK INPUT ")
    for col in range(col_start_input, col_end_input):
        """check cont"""
        if str(sheet_TCs.cell(Row_Type, col).value) == 'cont':
            """Check Tolerance YES/NO"""
            if sheet_TCs.cell(Row_Tolerance, col).value is not None:
                value_tol = float(sheet_TCs.cell(Row_Tolerance, col).value)
            else:
                value_tol = 'None'
                print_error(" ====> Error: Thieu Tolerance", col)

            """Check must be out max"""
            flag_ok_max = E_NOT_OK
            flag_ok_min = E_NOT_OK
            flag_not_out_max = E_NOT_OK
            flag_not_out_min = E_NOT_OK
            flag_not_mid_value = E_NOT_OK
            value_max = 0
            value_min = 1
            if sheet_TCs.cell(Row_Max, col).value is not None:
                value_max = float(sheet_TCs.cell(Row_Max, col).value)
                flag_ok_max = E_OK
            else:
                flag_ok_max = E_NOT_OK

            if sheet_TCs.cell(Row_Min, col).value is not None:
                value_min = float(sheet_TCs.cell(Row_Min, col).value)
                flag_ok_min = E_OK
            else:
                flag_ok_min = E_NOT_OK

            if flag_ok_min == E_OK and flag_ok_max == E_OK:
                '''Check input out max'''
                for row in range(Row_TC1, max_row_table + 1):
                    if sheet_TCs.cell(row, col).value is not None:
                        if float(sheet_TCs.cell(row, col).value) > value_max:
                            flag_not_out_max = E_NOT_OK
                            break
                        else:
                            flag_not_out_max = E_OK
                    else:
                        break

                '''Check input out min'''
                for row in range(Row_TC1, max_row_table + 1):
                    if sheet_TCs.cell(row, col).value is not None:
                        if float(sheet_TCs.cell(row, col).value) < value_min:
                            flag_not_out_min = E_NOT_OK
                            break
                        else:
                            flag_not_out_min = E_OK
                    else:
                        break

                '''Check input not mid value'''
                for row in range(Row_TC1, max_row_table + 1):
                    if sheet_TCs.cell(row, col).value is not None:
                        if float(sheet_TCs.cell(row, col).value) < value_max and float(
                                sheet_TCs.cell(row, col).value) > value_min:
                            flag_not_mid_value = E_NOT_OK
                            break
                        else:
                            flag_not_mid_value = E_OK
                    else:
                        break
            else:
                print_error(" ====> Error: Missing row value Max/Min", col)

            '''print resuit'''
            if flag_not_out_max == E_OK:
                print_error(" ====> Error: None out max: ", col)
            if flag_not_out_min == E_OK:
                print_error(" ====> Error: None out min: ", col)
            if flag_not_mid_value == E_OK:
                print_error(" ====> Error: None mid value: ", col)

        """check log"""
        if str(sheet_TCs.cell(Row_Type, col).value) == 'log':
            """Check Tolerance YES/NO and must be 0 or 1"""
            if sheet_TCs.cell(Row_Tolerance, col).value is not None:
                if isinstance(sheet_TCs.cell(Row_Tolerance, col).value, int):
                    value_tol = int(sheet_TCs.cell(Row_Tolerance, col).value)
                else:
                    print_error(" ====> Error: Row Tolerance must be 0 or 1", col)
            else:
                value_tol = 'None'
                print_error(" ====> Error: Thieu Tolerance", col)

            flag_true = E_NOT_OK
            "Check must be has TRUE & FALSE"
            flag_false = E_NOT_OK
            "Check must be has TRUE & FALSE"
            flag_format = E_OK
            "Check format TRUE/FALSE, not 1/0"
            for row in range(Row_TC1, max_row_table + 1):
                if sheet_TCs.cell(row, col).value is not None:
                    if str(sheet_TCs.cell(row, col).value) == 'True':
                        flag_true = E_OK
                    if str(sheet_TCs.cell(row, col).value) == 'False':
                        flag_false = E_OK
                else:
                    break

            for row in range(Row_TC1, max_row_table + 1):
                if sheet_TCs.cell(row, col).value is not None:
                    if str(sheet_TCs.cell(row, col).value) != 'True' and str(
                            sheet_TCs.cell(row, col).value) != 'False':
                        flag_format = E_NOT_OK
                        break

            if flag_true == E_NOT_OK or flag_false == E_NOT_OK:
                print_error(" ====> Error: Thieu TRUE/FALSE: ", col)
            if flag_format == E_NOT_OK:
                print_error(" ====> Error: Wrong format Bool  ", col)

        '''Check input enum'''
        if str(sheet_TCs.cell(Row_Type, col).value) == 'enum':
            """Check Tolerance YES/NO and must be int"""
            if sheet_TCs.cell(Row_Tolerance, col).value is not None:
                value_tol = float(sheet_TCs.cell(Row_Tolerance, col).value)
                if value_tol == 0 or value_tol == 1:
                    pass
                else:
                    print_error(" ====> Error: Row Tolerance must be 'int' ", col)

                """Check max enum"""
                flag_ok_max = E_NOT_OK
                flag_ok_min = E_NOT_OK
                flag_enum_max = E_NOT_OK
                flag_enum_min = E_NOT_OK
                flag_enum_mid = E_NOT_OK
                flag_enum_out_max = E_OK
                flag_enum_format = E_OK
                value_max = 0
                value_min = 0
                number_of_enum = 0
                if sheet_TCs.cell(Row_Max, col).value is not None:
                    value_max = float(sheet_TCs.cell(Row_Max, col).value)
                    flag_ok_max = E_OK
                else:
                    flag_ok_max = E_NOT_OK

                if sheet_TCs.cell(Row_Min, col).value is not None:
                    value_min = float(sheet_TCs.cell(Row_Min, col).value)
                    flag_ok_min = E_OK
                else:
                    flag_ok_min = E_NOT_OK

                if flag_ok_min == E_OK and flag_ok_max == E_OK:
                    '''Check enum max'''
                    " quet theo hang de in name enum "
                    for row in range(Row_TC1, max_row_table + 1):
                        if sheet_TCs.cell(row, col).value is not None:
                            for col_enum in range(TC_No, max_column_table + 1):
                                if isinstance(sheet_TCs.cell(row, col).value, str):
                                    name_enum = sheet_TCs.cell(row, col).value
                                    if str(sheet_TCs.cell(Row_Enum, col_enum).value) == name_enum:
                                        number_of_enum = float(sheet_TCs.cell(Row_Enum + 1, col_enum).value)
                                        break
                                else:
                                    """Neu dinh dang sai thi bat co bao"""
                                    flag_enum_format = E_NOT_OK

                            if number_of_enum == value_max:
                                flag_enum_max = E_OK
                            if number_of_enum == value_min:
                                flag_enum_min = E_OK
                            if value_max > number_of_enum > value_min:
                                flag_enum_mid = E_OK
                            # cover for max_enum = 1; min_enum = 0;
                            if value_max - value_min == 1:   # for enum                   
                                flag_enum_mid = E_OK
                            if number_of_enum > value_max:
                                flag_enum_out_max = E_NOT_OK

                        else:
                            break

                    if flag_enum_max == E_NOT_OK:
                        print_error(" ====> Error: Missing Enum max", col)

                    if flag_enum_min == E_NOT_OK:
                        print_error(" ====> Error: Missing Enum min", col)

                    if flag_enum_mid == E_NOT_OK:
                        print_error(" ====> Error: Missing Enum mid", col)

                    if flag_enum_format == E_NOT_OK:
                        print_error(" ====> Error: Wrong format Enum - Not is number ", col)

                    if flag_enum_out_max == E_NOT_OK:
                        print_error(" ====> Error: Out range max Enum ", col)

            else:
                value_tol = 'None'
                print_error(" ====> Error: Thieu Tolerance", col)

    """END check_input"""
    return
# End def check_input():


def check_as_input():
    """
    Check xem sheet nay co LOCAL VARIABLES AS INPUT hay khong?
    Neu co:
        cont
            Check Tolerance YES/NO
            Check must not be out max
            Check must not be out min
            Check must be mid value
        log
            Check Tolerance 0:1
            Check must be has TRUE & FALSE
            Check fomat "TRUE/FALSE", not 1/0
        enum
            Print warning
            Check Tolerance
    """
    flag_yes_as_input = E_NOT_OK
    "Flag check xem sheet co LOCAL VARIABLES AS INPUT hay khong?"
    col_start_as_input = 0
    col_end_as_input = 0
    col_start_outputs = 0
    # Tim cot LOCAL VARIABLES AS INPUT
    for col in range(1, max_column_table + 1):
        if str(sheet_TCs.cell(Row_Title, col).value) == 'LOCAL VARIABLES AS INPUT':
            flag_yes_as_input = E_OK
            col_start_as_input = col

    # Tim cot OUTPUTS
    for col in range(1, max_column_table + 1):
        if str(sheet_TCs.cell(Row_Title, col).value) == 'OUTPUTS':
            col_start_outputs = col
            break
    # neu cot as input nam ngoai outputs thi bao error
    if col_start_outputs < col_start_as_input:
        print_error(" ====> Error: Check location for LOCAL VARIABLES AS INPUT ")

    if flag_yes_as_input == E_OK and col_start_outputs > col_start_as_input:
        print_notice("======== CHECK AS INPUT ")
        for col in range(col_start_as_input + 1, max_column_table + 1):
            if sheet_TCs.cell(Row_Title, col).value is not None:
                col_end_as_input = col
                break
        "Check as input"
        for col in range(col_start_as_input, col_end_as_input):
            # check Missing Row 'Type'
            if sheet_TCs.cell(Row_Type, col).value is None:
                print_error(" ====> Error: Missing Row 'Type' ", col)

            "Check cont"
            if str(sheet_TCs.cell(Row_Type, col).value) == 'cont':
                """Check Tolerance YES/NO"""
                if sheet_TCs.cell(Row_Tolerance, col).value is not None:
                    value_tol = float(sheet_TCs.cell(Row_Tolerance, col).value)
                else:
                    value_tol = 'None'
                    print_error(" ====> Error: Thieu Tolerance", col)

                """Check must be out max"""
                flag_ok_max = E_NOT_OK
                flag_ok_min = E_NOT_OK
                flag_max = E_NOT_OK
                flag_out_max = E_OK
                flag_min = E_NOT_OK
                flag_out_min = E_OK
                flag_mid_value = E_NOT_OK
                value_max = 0
                value_min = 0
                if sheet_TCs.cell(Row_Max, col).value is not None:
                    value_max = float(sheet_TCs.cell(Row_Max, col).value)
                    flag_ok_max = E_OK
                else:
                    flag_ok_max = E_NOT_OK

                if sheet_TCs.cell(Row_Min, col).value is not None:
                    value_min = float(sheet_TCs.cell(Row_Min, col).value)
                    flag_ok_min = E_OK
                else:
                    flag_ok_min = E_NOT_OK

                if flag_ok_min == E_OK and flag_ok_max == E_OK:
                    '''Check input max'''
                    for row in range(Row_TC1, max_row_table + 1):
                        if sheet_TCs.cell(row, col).value is not None:
                            if float(sheet_TCs.cell(row, col).value) == value_max:
                                flag_max = E_OK

                            if float(sheet_TCs.cell(row, col).value) > value_max:
                                flag_out_max = E_NOT_OK
                        else:
                            break

                    '''Check input min'''
                    for row in range(Row_TC1, max_row_table + 1):
                        if sheet_TCs.cell(row, col).value is not None:
                            if float(sheet_TCs.cell(row, col).value) == value_min:
                                flag_min = E_OK

                            if float(sheet_TCs.cell(row, col).value) < value_min:
                                flag_out_min = E_NOT_OK
                        else:
                            break

                    '''Check input not mid value'''
                    for row in range(Row_TC1, max_row_table + 1):
                        if sheet_TCs.cell(row, col).value is not None:
                            if float(sheet_TCs.cell(row, col).value) < value_max and float(
                                    sheet_TCs.cell(row, col).value) > value_min:
                                flag_mid_value = E_OK
                                break
                            else:
                                flag_mid_value = E_NOT_OK
                        else:
                            break

                    '''print resuit'''
                    if flag_max == E_NOT_OK:
                        print_error(" ====> Error: None value max: ", col)
                    if flag_min == E_NOT_OK:
                        print_error(" ====> Error: None value min: ", col)
                    if flag_mid_value == E_NOT_OK:
                        print_error(" ====> Error: None mid value: ", col)
                    if flag_out_max == E_NOT_OK:
                        print_error(" ====> Error: Out range max: ", col)
                    if flag_out_min == E_NOT_OK:
                        print_error(" ====> Error: Out range min: ", col)
                else:
                    print_error(" ====> Error: Missing row value Max/Min", col)

            "Check log"
            if str(sheet_TCs.cell(Row_Type, col).value) == 'log':
                """Check Tolerance YES/NO and must be 0 or 1"""
                if sheet_TCs.cell(Row_Tolerance, col).value is not None:
                    value_tol = float(sheet_TCs.cell(Row_Tolerance, col).value)
                    if value_tol == 0 or value_tol == 1:
                        pass
                    else:
                        print_error(" ====> Error: Row Tolerance must be 0 or 1", col)
                else:
                    value_tol = 'None'
                    print_error(" ====> Error: Thieu Tolerance", col)

                flag_true = 0
                "Check must be has TRUE & FALSE"
                flag_false = 0
                "Check must be has TRUE & FALSE"
                flag_format = 0
                "Check format TRUE/FALSE, not 1/0"
                for row in range(Row_TC1, max_row_table + 1):
                    if sheet_TCs.cell(row, col).value is not None:
                        if str(sheet_TCs.cell(row, col).value) == 'True':
                            flag_true = 1
                        if str(sheet_TCs.cell(row, col).value) == 'False':
                            flag_false = 1
                    else:
                        break

                for row in range(Row_TC1, max_row_table + 1):
                    if sheet_TCs.cell(row, col).value is not None:
                        if str(sheet_TCs.cell(row, col).value) != 'True' and str(
                                sheet_TCs.cell(row, col).value) != 'False':
                            flag_format = 1
                            break

                if flag_true == 0 or flag_false == 0:
                    print_error(" ====> Error: Thieu TRUE/FALSE: ", col)
                if flag_format == 1:
                    print_error(" ====> Error: Wrong format Bool  ", col)
            '''Check input enum'''
            if str(sheet_TCs.cell(Row_Type, col).value) == 'enum':
                """Check Tolerance YES/NO and must be int"""
                if sheet_TCs.cell(Row_Tolerance, col).value is not None:
                    if isinstance(sheet_TCs.cell(Row_Tolerance, col).value, int):
                        value_tol = int(sheet_TCs.cell(Row_Tolerance, col).value)
                    else:
                        print_error(" ====> Error: Row Tolerance must be 'int' ", col)

                    """Check max enum"""
                    flag_ok_max = 0
                    flag_ok_min = 0
                    flag_enum_max = 0
                    flag_enum_min = 0
                    flag_enum_mid = 0
                    flag_enum_out_max = E_OK
                    flag_enum_format = 0
                    value_max = 0
                    value_min = 0
                    number_of_enum = 0
                    if sheet_TCs.cell(Row_Max, col).value is not None:
                        value_max = float(sheet_TCs.cell(Row_Max, col).value)
                        flag_ok_max = 1
                    else:
                        flag_ok_max = 0

                    if sheet_TCs.cell(Row_Min, col).value is not None:
                        value_min = float(sheet_TCs.cell(Row_Min, col).value)
                        flag_ok_min = 1
                    else:
                        flag_ok_min = 0
                    '''Check enum max min mid'''
                    if flag_ok_min == 1 and flag_ok_max == 1:
                        for row in range(Row_TC1, max_row_table + 1):
                            if sheet_TCs.cell(row, col).value is not None:
                                for col_enum in range(TC_No, max_column_table + 1):
                                    if isinstance(sheet_TCs.cell(row, col).value, str):
                                        name_enum = sheet_TCs.cell(row, col).value
                                        if str(sheet_TCs.cell(Row_Enum, col_enum).value) == name_enum:
                                            number_of_enum = float(sheet_TCs.cell(Row_Enum + 1, col_enum).value)
                                            break
                                    else:
                                        """Neu dinh dang sai thi bat co bao"""
                                        flag_enum_format = 1

                                if number_of_enum == value_max:
                                    flag_enum_max = 1
                                if number_of_enum == value_min:
                                    flag_enum_min = 1
                                if value_max > number_of_enum > value_min:
                                    if value_max - value_min != 1:
                                        # cover for max_enum = 1; min_enum = 0;
                                        flag_enum_mid = E_OK
                                if number_of_enum > value_max:
                                    flag_enum_out_max = E_NOT_OK
                            else:
                                break

                        if flag_enum_max == 0:
                            print_error(" ====> Error: Missing Enum max", col)
                        if flag_enum_min == 0:
                            print_error(" ====> Error: Missing Enum min", col)
                        if flag_enum_mid == 0:
                            print_error(" ====> Error: Missing Enum mid", col)
                        if flag_enum_format == 1:
                            print_error(" ====> Error: Wrong format Enum - Not is number ", col)

                        if flag_enum_out_max == E_NOT_OK:
                            print_error(" ====> Error: Out range max Enum ", col)
                else:
                    value_tol = 'None'
                    print_error(" ====> Error: Thieu Tolerance", col)

    """END check_as_input"""
    return
# End def check_as_input():


def check_imported_parameters():
    """
    Check xem co IMPORTED PARAMETERS hay khong?
    Neu co:
        cont
            Check Tolerance
            inf -inf number
            Check must not be out max
            Check must not be out min
            Check must be mid value
            Check when value is const
        log
            Check Tolerance 0:1
            Check must be has TRUE & FALSE
            Check fomat "TRUE/FALSE", not 1/0
        enum
            Print warning
            Check none Tolerance
    :return:
    """
    flag_yes_imp_parm = 0
    "Flag check xem co IMPORTED PARAMETERS hay khong"
    col_start_imp_parm = 0
    col_end_imp_parm = 0
    for col in range(1, max_column_table + 1):
        if str(sheet_TCs.cell(Row_Title, col).value) == 'IMPORTED PARAMETERS':
            flag_yes_imp_parm = 1
            col_start_imp_parm = col
            break
    if flag_yes_imp_parm == 1:
        print_notice("======== CHECK IMPORTED PARAMETERS ")
        for col in range(col_start_imp_parm + 1, max_column_table + 1):
            if sheet_TCs.cell(Row_Title, col).value is not None:
                col_end_imp_parm = col
                break
        '''Check IMPORTED PARAMETERS '''
        for col in range(col_start_imp_parm, col_end_imp_parm):
            # '''Check IMPORTED PARAMETERS cont'''
            if sheet_TCs.cell(Row_Type, col).value is None:
                print_error(" ====> Error: Missing Row 'Type' ", col)

            # """Check Tolerance YES/NO"""
            if sheet_TCs.cell(Row_Tolerance, col).value is not None:
                print_error(" ====> Error: Khong can Tolerance", col)
            else:
                value_tol = 'None'
            #  Check 'cont'
            if str(sheet_TCs.cell(Row_Type, col).value) == 'cont':
                if str(sheet_TCs.cell(Row_Min, col).value) == '-inf':
                    pass
                elif str(sheet_TCs.cell(Row_Min, col).value) == '-INF':
                    pass
                elif str(sheet_TCs.cell(Row_Min, col).value) == '-Inf':
                    pass
                else:

                    
                    """Check must be out max"""
                    flag_ok_max = 0
                    flag_ok_min = 0
                    flag_max = 0
                    flag_out_max = 0
                    flag_min = 0
                    flag_out_min = 0
                    flag_mid_value = 0
                    value_max = 0
                    value_min = 0
                    if sheet_TCs.cell(Row_Max, col).value is not None:
                        value_max = float(sheet_TCs.cell(Row_Max, col).value)
                        flag_ok_max = 1
                    else:
                        flag_ok_max = 0

                    if sheet_TCs.cell(Row_Min, col).value is not None:
                        value_min = float(sheet_TCs.cell(Row_Min, col).value)
                        flag_ok_min = 1
                    else:
                        flag_ok_min = 0

                    if flag_ok_max == 1 and flag_ok_min == 1:
                        '''Check input max'''
                        for row in range(Row_TC1, max_row_table + 1):
                            if sheet_TCs.cell(row, col).value is not None:
                                if float(sheet_TCs.cell(row, col).value) == value_max:
                                    flag_max = 1

                                if float(sheet_TCs.cell(row, col).value) > value_max:
                                    flag_out_max = 1
                            else:
                                break

                        '''Check input min'''
                        for row in range(Row_TC1, max_row_table + 1):
                            if sheet_TCs.cell(row, col).value is not None:
                                if float(sheet_TCs.cell(row, col).value) == value_min:
                                    flag_min = 1

                                if float(sheet_TCs.cell(row, col).value) < value_min:
                                    flag_out_min = 1
                            else:
                                break

                        '''Check input not mid value'''
                        for row in range(Row_TC1, max_row_table + 1):
                            if sheet_TCs.cell(row, col).value is not None:
                                if float(sheet_TCs.cell(row, col).value) < value_max and float(
                                        sheet_TCs.cell(row, col).value) > value_min:
                                    flag_mid_value = 0
                                    break
                                else:
                                    flag_mid_value = 1
                            else:
                                break

                        '''Check xem co phai const hay khong? '''
                        if value_max == value_min:# or value_max - value_min == 1:
                            flag_mid_value = 0

                        '''print resuit'''
                        if flag_max == 0:
                            print_error(" ====> Error: None value max: ", col)
                        if flag_min == 0:
                            print_error(" ====> Error: None value min: ", col)
                        if flag_mid_value == 1:
                            print_error(" ====> Error: None mid value: ", col)
                        if flag_out_max == 1:
                            print_error(" ====> Error: Out range max: ", col)
                        if flag_out_min == 1:
                            print_error(" ====> Error: Out range min: ", col)
                    else:
                        print_error(" ====> Error: Missing row value Max/Min", col)

            # "Check log"
            if str(sheet_TCs.cell(Row_Type, col).value) == 'log':

                flag_true = 0
                "Check must be has TRUE & FALSE"
                flag_false = 0
                "Check must be has TRUE & FALSE"
                flag_format = 0
                "Check format TRUE/FALSE, not 1/0"
                for row in range(Row_TC1, max_row_table + 1):
                    if sheet_TCs.cell(row, col).value is not None:
                        if str(sheet_TCs.cell(row, col).value) == 'True':
                            flag_true = 1
                        if str(sheet_TCs.cell(row, col).value) == 'False':
                            flag_false = 1
                    else:
                        break

                for row in range(Row_TC1, max_row_table + 1):
                    if sheet_TCs.cell(row, col).value is not None:
                        if str(sheet_TCs.cell(row, col).value) != 'True' and str(
                                sheet_TCs.cell(row, col).value) != 'False':
                            flag_format = 1
                            break

                if flag_true == 0 or flag_false == 0:
                    print_error(" ====> Error: Thieu TRUE/FALSE: ", col)
                if flag_format == 1:
                    print_error(" ====> Error: Wrong format Bool  ", col)
            
            # '''Check input enum'''
            if str(sheet_TCs.cell(Row_Type, col).value) == 'enum':
                """Check max enum"""
                flag_ok_max = 0
                flag_ok_min = 0
                flag_enum_max = 0
                flag_enum_min = 0
                flag_enum_mid = 0
                flag_enum_out_max = E_OK
                flag_enum_format = 0
                value_max = 0
                value_min = 0
                number_of_enum = 0
                if sheet_TCs.cell(Row_Max, col).value is not None:
                    value_max = float(sheet_TCs.cell(Row_Max, col).value)
                    flag_ok_max = 1
                else:
                    flag_ok_max = 0

                if sheet_TCs.cell(Row_Min, col).value is not None:
                    value_min = float(sheet_TCs.cell(Row_Min, col).value)
                    flag_ok_min = 1
                else:
                    flag_ok_min = 0
                '''Check enum max min mid'''
                if flag_ok_min == 1 and flag_ok_max == 1:
                    for row in range(Row_TC1, max_row_table + 1):
                        if sheet_TCs.cell(row, col).value is not None:
                            for col_enum in range(TC_No, max_column_table + 1):
                                if isinstance(sheet_TCs.cell(row, col).value, str):
                                    name_enum = sheet_TCs.cell(row, col).value
                                    if str(sheet_TCs.cell(Row_Enum, col_enum).value) == name_enum:
                                        number_of_enum = float(sheet_TCs.cell(Row_Enum + 1, col_enum).value)
                                        break
                                else:
                                    """Neu dinh dang sai thi bat co bao"""
                                    flag_enum_format = 1

                            if number_of_enum == value_max:
                                flag_enum_max = 1
                            if number_of_enum == value_min:
                                flag_enum_min = 1
                            if value_max > number_of_enum > value_min:
                                flag_enum_mid = E_OK
                            # cover for max_enum = 1; min_enum = 0;
                            if value_max - value_min == 1:                               
                                flag_enum_mid = E_OK

                            if number_of_enum > value_max:
                                flag_enum_out_max = E_NOT_OK    
                        else:
                            break

                    if flag_enum_max == 0:
                        print_error(" ====> Error: Missing Enum max", col)
                    if flag_enum_min == 0:
                        print_error(" ====> Error: Missing Enum min", col)
                    if flag_enum_mid == 0:
                        print_error(" ====> Error: Missing Enum mid", col)
                    if flag_enum_format == 1:
                        print_error(" ====> Error: Wrong format Enum - Not is number ", col)
                    
                    if flag_enum_out_max == E_NOT_OK:
                        print_error(" ====> Error: Out range max Enum ", col)


    """END check IMPORTED PARAMETERS"""
    return
# End def check_imported_parameters():


def check_local_variable():
    """
    Check xem co LOCAL VARIABLES hay khong?
    Neu co:
        cont
            Check Tolenrance
            Check must not be out max
            Check must not be out min
            Check must be mid value
    :return:
    """
    flag_yes_lcal_var = 0
    '''Flag check xem co LOCAL VARIABLES hay khong?'''
    col_start_lcal_var = 0
    col_end_lcal_var = 0
    for col in range(1, max_column_table + 1):
        if str(sheet_TCs.cell(Row_Title, col).value) == 'LOCAL VARIABLES':
            flag_yes_lcal_var = 1
            col_start_lcal_var = col
            break
    if flag_yes_lcal_var == 1:
        print_notice("======== CHECK LOCAL VARIABLES ")
        for col in range(col_start_lcal_var + 1, max_column_table + 1):
            if sheet_TCs.cell(Row_Title, col).value is not None:
                col_end_lcal_var = col
                break
        '''Check LOCAL VARIABLES'''
        for col in range(col_start_lcal_var, col_end_lcal_var):
            '''check cont'''
            if str(sheet_TCs.cell(Row_Type, col).value) == 'cont':
                """Check Tolerance YES/NO"""
                if sheet_TCs.cell(Row_Tolerance, col).value is None:
                    print_error(" ====> Error: Thieu Tolerance", col)

                """Check must be out max"""
                flag_ok_max = 0
                flag_ok_min = 0
                flag_max = 0
                flag_out_max = 0
                flag_min = 0
                flag_out_min = 0
                flag_mid_value = 0
                flag_check_round = 0
                flag_div0 = 0
                value_max = 0
                value_min = 0
                if sheet_TCs.cell(Row_Max, col).value is not None:
                    value_max = float(sheet_TCs.cell(Row_Max, col).value)
                    flag_ok_max = 1
                else:
                    flag_ok_max = 0

                if sheet_TCs.cell(Row_Min, col).value is not None:
                    value_min = float(sheet_TCs.cell(Row_Min, col).value)
                    flag_ok_min = 1
                else:
                    flag_ok_min = 0

                if flag_ok_min == 1 and flag_ok_max == 1:
                    '''Check input max'''
                    for row in range(Row_TC1, max_row_table + 1):
                        if sheet_TCs.cell(row, col).value is not None:
                            try:
                                if float(sheet_TCs.cell(row, col).value) == value_max:
                                    flag_max = 1

                                if float(sheet_TCs.cell(row, col).value) > value_max:
                                    flag_out_max = 1
                            except ValueError:
                                flag_div0 = 1
                        else:
                            break

                    '''Check input min'''
                    for row in range(Row_TC1, max_row_table + 1):
                        if sheet_TCs.cell(row, col).value is not None:
                            try:
                                if float(sheet_TCs.cell(row, col).value) == value_min:
                                    flag_min = 1

                                if float(sheet_TCs.cell(row, col).value) < value_min:
                                    flag_out_min = 1
                            except ValueError:
                                flag_div0 = 1
                        else:
                            break

                    '''Check input not mid value'''
                    for row in range(Row_TC1, max_row_table + 1):
                        if sheet_TCs.cell(row, col).value is not None:
                            try:
                                if float(sheet_TCs.cell(row, col).value) < value_max and float(
                                        sheet_TCs.cell(row, col).value) > value_min:
                                    flag_mid_value = 0  # co mid value
                                    break
                                else:
                                    flag_mid_value = 1  # khong co mid value
                            except ValueError:
                                flag_div0 = 1
                        else:
                            break

                    '''Check xem khong co mid value co phu hop hay khong?'''
                    if flag_mid_value == 1:  # khong co mid value
                        for row in range(Row_TC1, max_row_table + 1):
                            if sheet_TCs.cell(row, col).value is not None:
                                try:
                                    if float(sheet_TCs.cell(row, col).value) == value_max:
                                        flag_check_round = 0
                                    elif float(sheet_TCs.cell(row, col).value) == value_min:
                                        flag_check_round = 0
                                    else:
                                        flag_check_round = 1
                                        break

                                except ValueError:
                                    flag_div0 = 1
                            else:
                                break

                    '''print resuit'''
                    if flag_max == 0:
                        print_error(" ====> Error: None value max: ", col)
                    if flag_min == 0:
                        print_error(" ====> Error: None value min: ", col)
                    if flag_check_round == 1:
                        print_error(" ====> Error: Check round min max : ", col)
                    if flag_out_max == 1:
                        print_error(" ====> Error: Out range max: ", col)
                    if flag_out_min == 1:
                        print_error(" ====> Error: Out range min: ", col)
                    if flag_div0 == 1:
                        print_error(" ====> Error: Must not be 'DIV/0' : ", col)

                else:
                    print_error(" ====> Error: Missing row value Max/Min", col)
    '''END check local param'''
    return
# End def check_local_variable():


def check_parameters():
    """
    Check xem co PARAMETERS hay khong?
    Neu co:
        cont
            Check Tolenrance
            Check must not be out max
            Check must not be out min
            Check must be mid value
        log
            Check Tolerance 0:1
            Check must be has TRUE & FALSE
            Check fomat "TRUE/FALSE", not 1/0
    :return:
    """
    flag_yes_parameters = 0
    '''Flag check xem co PARAMETERS hay khong? '''
    col_start_parameters = 0
    col_end_parameters = 0
    for col in range(1, max_column_table + 1):
        if str(sheet_TCs.cell(Row_Title, col).value) == 'PARAMETERS':
            print_notice("======== CHECK PARAMETERS ")
            flag_yes_parameters = 1
            col_start_parameters = col
            break
    if flag_yes_parameters == 1:
        for col in range(col_start_parameters + 1, max_column_table + 1):
            if sheet_TCs.cell(Row_Title, col).value is not None:
                col_end_parameters = col
                break
        '''Check PARAMETERS'''
        for col in range(col_start_parameters, col_end_parameters):
            '''Check cont'''
            if str(sheet_TCs.cell(Row_Type, col).value) == 'cont':
                """Check Tolerance YES/NO"""
                if sheet_TCs.cell(Row_Tolerance, col).value is not None:
                    value_tol = float(sheet_TCs.cell(Row_Tolerance, col).value)
                else:
                    value_tol = 'None'
                    print_error(" ====> Error: Thieu Tolerance", col)

                """Check must be out max"""
                flag_ok_max = 0
                flag_ok_min = 0
                flag_out_max = 0
                flag_out_min = 0
                flag_div0 = 0
                value_max = 0
                value_min = 0
                if sheet_TCs.cell(Row_Max, col).value is not None:
                    value_max = float(sheet_TCs.cell(Row_Max, col).value)
                    flag_ok_max = 1
                else:
                    flag_ok_max = 0

                if sheet_TCs.cell(Row_Min, col).value is not None:
                    value_min = float(sheet_TCs.cell(Row_Min, col).value)
                    flag_ok_min = 1
                else:
                    flag_ok_min = 0

                if flag_ok_min == 1 and flag_ok_max == 1:
                    '''Check input max'''
                    for row in range(Row_TC1, max_row_table + 1):
                        if sheet_TCs.cell(row, col).value is not None:
                            try:
                                if float(sheet_TCs.cell(row, col).value) > value_max:
                                    flag_out_max = 1
                            except ValueError:
                                flag_div0 = 1
                        else:
                            break

                    '''Check input min'''
                    for row in range(Row_TC1, max_row_table + 1):
                        if sheet_TCs.cell(row, col).value is not None:
                            try:
                                if float(sheet_TCs.cell(row, col).value) < value_min:
                                    flag_out_min = 1
                            except ValueError:
                                flag_div0 = 1
                        else:
                            break

                    '''print resuit'''
                    if flag_out_max == 1:
                        print_error(" ====> Error: Out range max: ", col)
                    if flag_out_min == 1:
                        print_error(" ====> Error: Out range min: ", col)
                    if flag_div0 == 1:
                        print_warning(" ====> WARNING: DIV/0: ", col)

                else:
                    print_error(" ====> Error: Missing row value Max/Min", col)
            '''Check log'''
            if str(sheet_TCs.cell(Row_Type, col).value) == 'log':
                """Check Tolerance YES/NO and must be 0 or 1"""
                if sheet_TCs.cell(Row_Tolerance, col).value is not None:
                    value_tol = float(sheet_TCs.cell(Row_Tolerance, col).value)
                    if value_tol == 0 or value_tol == 1:
                        pass
                    else:
                        print_error(" ====> Error: Row Tolerance must be 0 or so nguyen ", col)
                else:
                    value_tol = 'None'
                    print_error(" ====> Error: Thieu Tolerance", col)

                flag_true = 0
                "Check must be has TRUE & FALSE"
                flag_false = 0
                "Check must be has TRUE & FALSE"
                flag_format = 0
                "Check format TRUE/FALSE, not 1/0"
                for row in range(Row_TC1, max_row_table + 1):
                    if sheet_TCs.cell(row, col).value is not None:
                        if str(sheet_TCs.cell(row, col).value) == 'True':
                            flag_true = 1
                        if str(sheet_TCs.cell(row, col).value) == 'False':
                            flag_false = 1
                    else:
                        break

                for row in range(Row_TC1, max_row_table + 1):
                    if sheet_TCs.cell(row, col).value is not None:
                        if str(sheet_TCs.cell(row, col).value) != 'True' and str(
                                sheet_TCs.cell(row, col).value) != 'False':
                            flag_format = 1
                            break

                if flag_true == 0 and flag_false == 0:
                    print_error(" ====> Error: Thieu TRUE/FALSE: ", col)
                if flag_format == 1:
                    print_error(" ====> Error: Wrong format Bool  ", col)
    '''END check PARAMETERS'''
    return
# End def check_parameters():


def check_output():
    """
    cont
        Check Tolerance
        Check must not be out max
        Check must not be out min
        Check must be mid value
        Check when value is const
    log
        Check Tolerance 0:1
        Check must be has TRUE & FALSE
        Check fomat "TRUE/FALSE", not 1/0
    enum
        Check Tolerance
        Check enum not type number
    :return:
    """

    print_notice("======== CHECK OUTPUTS ")
    col_start_outputs = 0
    col_end_outputs = 0
    for col in range(1, max_column_table + 1):
        if str(sheet_TCs.cell(Row_Title, col).value) == 'OUTPUTS':
            col_start_outputs = col
            break

    for col in range(col_start_outputs + 1, max_column_table + 1):
        if sheet_TCs.cell(Row_Title, col).value is not None:
            col_end_outputs = col
            break
    '''Check OUTPUT'''
    for col in range(col_start_outputs, col_end_outputs):
        '''Check cont'''
        if str(sheet_TCs.cell(Row_Type, col).value) == 'cont':
            """Check Tolerance YES/NO"""
            if sheet_TCs.cell(Row_Tolerance, col).value is not None:
                value_tol = float(sheet_TCs.cell(Row_Tolerance, col).value)
            else:
                value_tol = 'None'
                print_error(" ====> Error: Thieu Tolerance", col)

            """Check must be out max"""
            flag_ok_max = 0
            flag_ok_min = 0
            flag_out_max = 0
            flag_out_min = 0
            flag_div0 = 0
            value_max = 0
            value_min = 0
            if sheet_TCs.cell(Row_Max, col).value is not None:
                value_max = float(sheet_TCs.cell(Row_Max, col).value)
                flag_ok_max = 1
            else:
                flag_ok_max = 0

            if sheet_TCs.cell(Row_Min, col).value is not None:
                value_min = float(sheet_TCs.cell(Row_Min, col).value)
                flag_ok_min = 1
            else:
                flag_ok_min = 0

            if flag_ok_min == 1 and flag_ok_max == 1:
                '''Check input max'''
                for row in range(Row_TC1, max_row_table + 1):
                    if sheet_TCs.cell(row, col).value is not None:
                        try:
                            if float(sheet_TCs.cell(row, col).value) > value_max:
                                flag_out_max = 1
                        except ValueError:
                            flag_div0 = 1
                    else:
                        break

                '''Check input min'''
                for row in range(Row_TC1, max_row_table + 1):
                    if sheet_TCs.cell(row, col).value is not None:
                        try:
                            if float(sheet_TCs.cell(row, col).value) < value_min:
                                flag_out_min = 1
                        except ValueError:
                            flag_div0 = 1
                    else:
                        break

                '''print resuit'''
                if flag_out_max == 1:
                    print_error(" ====> Error: Out range max: ", col)
                if flag_out_min == 1:
                    print_error(" ====> Error: Out range min: ", col)
                if flag_div0 == 1:
                    print_warning(" ====> WARNING: DIV/0: ", col)
            else:
                print_error(" ====> Error: Missing row value Max/Min", col)
        '''Check log'''
        if str(sheet_TCs.cell(Row_Type, col).value) == 'log':
            """Check Tolerance YES/NO and must be 0 or 1"""
            if sheet_TCs.cell(Row_Tolerance, col).value is not None:
                value_tol = float(sheet_TCs.cell(Row_Tolerance, col).value)
                if value_tol == 0 or value_tol == 1:
                    pass
                else:
                    print_error(" ====> Error: Row Tolerance must be 0 or 1", col)
            else:
                value_tol = 'None'
                print_error(" ====> Error: Thieu Tolerance", col)

            flag_true = 0
            "Check must be has TRUE or FALSE"
            flag_false = 0
            "Check must be has TRUE or FALSE"
            flag_format = 0
            "Check format TRUE/FALSE, not 1/0"
            for row in range(Row_TC1, max_row_table + 1):
                if sheet_TCs.cell(row, col).value is not None:
                    if str(sheet_TCs.cell(row, col).value) == 'True':
                        flag_true = 1
                    if str(sheet_TCs.cell(row, col).value) == 'False':
                        flag_false = 1
                else:
                    break

            for row in range(Row_TC1, max_row_table + 1):
                if sheet_TCs.cell(row, col).value is not None:
                    if str(sheet_TCs.cell(row, col).value) != 'True' and str(
                            sheet_TCs.cell(row, col).value) != 'False':
                        flag_format = 1
                        break

            if flag_true == 0 and flag_false == 0:
                print_error(" ====> Error: Thieu TRUE/FALSE: ", col)
            if flag_format == 1:
                print_error(" ====> Error: Wrong format Bool  ", col)
        '''Check input enum'''
        if str(sheet_TCs.cell(Row_Type, col).value) == 'enum':
            """Check Tolerance YES/NO and must be int"""
            if sheet_TCs.cell(Row_Tolerance, col).value is not None:
                if isinstance(sheet_TCs.cell(Row_Tolerance, col).value, int):
                    value_tol = int(sheet_TCs.cell(Row_Tolerance, col).value)
                else:
                    print_error(" ====> Error: Row Tolerance must be 'int' ", col)
                # check enum
                flag_ok_max = 0
                flag_ok_min = 0
                flag_enum_max = 0
                flag_enum_min = 0
                flag_enum_mid = 0
                flag_enum_out_max = E_OK
                flag_enum_format = 0
                value_max = 0
                value_min = 0
                number_of_enum = 0
                if sheet_TCs.cell(Row_Max, col).value is not None:
                    value_max = float(sheet_TCs.cell(Row_Max, col).value)
                    flag_ok_max = 1
                else:
                    flag_ok_max = 0

                if sheet_TCs.cell(Row_Min, col).value is not None:
                    value_min = float(sheet_TCs.cell(Row_Min, col).value)
                    flag_ok_min = 1
                else:
                    flag_ok_min = 0
                '''Check enum format'''
                if flag_ok_min == 1 and flag_ok_max == 1:
                    for row in range(Row_TC1, max_row_table + 1):
                        if sheet_TCs.cell(row, col).value is not None:
                            for col_enum in range(TC_No, max_column_table + 1):
                                if isinstance(sheet_TCs.cell(row, col).value, str):
                                    name_enum = sheet_TCs.cell(row, col).value
                                    if str(sheet_TCs.cell(Row_Enum, col_enum).value) == name_enum:
                                        number_of_enum = float(sheet_TCs.cell(Row_Enum + 1, col_enum).value)
                                        break
                                else:
                                    """Neu dinh dang sai thi bat co bao"""
                                    flag_enum_format = 1

                            # if number_of_enum == value_max:
                            #     flag_enum_max = 1
                            # if number_of_enum == value_min:
                            #     flag_enum_min = 1
                            # if value_max > number_of_enum > value_min:
                            #     flag_enum_mid = E_OK
                            # # cover for max_enum = 1; min_enum = 0;
                            # if value_max - value_min == 1:                               
                            #     flag_enum_mid = E_OK

                            if number_of_enum > value_max:
                                flag_enum_out_max = E_NOT_OK    
                        else:
                            break

                    # if flag_enum_max == 0:
                    #     print_error(" ====> Error: Missing Enum max", col)
                    # if flag_enum_min == 0:
                    #     print_error(" ====> Error: Missing Enum min", col)
                    # if flag_enum_mid == 0:
                    #     print_error(" ====> Error: Missing Enum mid", col)
                    if flag_enum_format == 1:
                        print_error(" ====> Error: Wrong format Enum - Not is number ", col)
                    
                    if flag_enum_out_max == E_NOT_OK:
                        print_error(" ====> Error: Out range max Enum ", col)

            else:
                value_tol = 'None'
                print_error(" ====> Error: Thieu Tolerance", col)
    '''End check output'''
    return
# End def check_output():


def check_descriptions():
    # "DESCRIPTIONS"
    for col in range(1, max_column_table + 1):
        if str(sheet_TCs.cell(Row_Title, col).value) == 'DESCRIPTIONS':
            for row in range(Row_TC1, max_row_table + 1):
                if sheet_TCs.cell(row, col).value is None:
                    print_notice("======== CHECK DESCRIPTIONS ")
                    print_error(" ====> Error: Missing Comment DESCRIPTIONS. ")
                    break
# End check DESCRIPTIONS


# ''' Neu co file '''
# """ Lay input tu bash shell """
try:
    # ''' Neu co file '''
    # """ Lay input tu bash shell """
    path_excel = str(sys.argv[1]) # link file TD
    try:
        # ''' Neu do dai link phu hop '''
        # Open file TD Excel
        wb = openpyxl.load_workbook(path_excel, data_only=True)
        # Khoi tao value name-Sheet can check
        sheet_Summary = wb['Summary']
        sheet_Rev_History = wb['Revision History']
    except:
        print_error(" ====> Error: Duong dan file dai. ")
        exit() # Exit khoi python

    try:
        value_RTRT_C0 = str(sys.argv[2]) # input chua data C0 RTRT
        value_RTRT_C1 = str(sys.argv[3]) # input chua data C1 RTRT
        value_RTRT_MCDC = str(sys.argv[4]) # input chua data MCDC RTRT
    except:    
        print_error(" ====> Error: No Data ReportRTRT.txt.")
        exit() # Exit khoi python
except:
    print_error(" ====> Error: Khong co file TD_*.xlsm. ")
    exit() # Exit khoi python


# check sheet revision_history
check_revision_history()
# chech MCDC C0C1
check_c0c1_mcdc()


# read tong so sheet TCs
Sheet_TC_End = len(wb.sheetnames)
# check sheet chua TCs
for num_sheet in range(Sheet_default, Sheet_TC_End):
    # khai bao bien va chon sheet
    try:
        sheet_TCs = wb.worksheets[num_sheet]
        if sheet_TCs.title == 'Testcases_Backup':
            continue       
        elif sheet_TCs.cell(Row_Title, TC_No).value != 'TC No.':
            continue
        elif sheet_TCs.cell(Row_TC1, TC_No).value is None:
            continue
    except:
        exit()

    # print name sheet dang test
    print_warning("---> Begin check Sheet: " + sheet_TCs.title)
    # khoi tao row col o sheet dang tesst    
    max_row_table = sheet_TCs.max_row
    max_column_table = sheet_TCs.max_column

    # """find number TCs"""
    for row in range(Row_TC1, max_row_table + 1):
        if sheet_TCs.cell(row, TC_No).value is None:
            max_row_table = row - 1 # break tai vi tri none
            break

    # """find location range col input"""
    col_start_input = 0
    for col in range(1, max_column_table + 1):
        if str(sheet_TCs.cell(Row_Title, col).value) == 'INPUTS':
            col_start_input = col
            break

    # """find location range col input"""
    for col in range(col_start_input + 1, max_column_table + 1):
        if sheet_TCs.cell(Row_Title, col).value is not None:
            col_end_input = col
            break


    check_TM_name()
    check_sum_TCs()
    check_value_row_max_min()
    check_input()
    check_as_input()
    check_imported_parameters()
    check_local_variable()
    check_parameters()
    check_output()
    check_descriptions()

    print_warning("---------> Check finished: " + sheet_TCs.title + '\n')
    # close file excel.
    wb.close()

